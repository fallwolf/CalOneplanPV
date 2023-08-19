import numpy as np
import xlrd
import matplotlib.pyplot as plt
import openpyxl
import math
import tkinter as tk
from tkinter import filedialog

filepath="d:/platedata/"
filename="30#-139.7"
workbook = xlrd.open_workbook(filepath+filename+'.xls')

# 选择第一个工作表
worksheet = workbook.sheet_by_index(0)
datalen=len(worksheet.col_values(2))

# 读取第2列数据--x
x = worksheet.col_values(2)[3:datalen]
ax = np.array(x)
# 读取第3列数据--y
y = worksheet.col_values(3)[3:datalen]
ay = np.array(y)
# 读取第3列数据--z
z = worksheet.col_values(4)[3:datalen]
az = np.array(z)
N = datalen-3
# ------------------------构建系数矩阵-----------------------------
A = np.array([[sum(ax ** 2), sum(ax * ay), sum(ax)],
              [sum(ax * ay), sum(ay ** 2), sum(ay)],
              [sum(ax), sum(ay), N]])

B = np.array([[sum(ax * az), sum(ay * az), sum(az)]])

# 求解
X = np.linalg.solve(A, B.T)

# 计算点到平面距离
pvbase=np.ones(len(ax))
i=0
while i<len(ax):
    pvbase[i]=(X[0]*ax[i]+X[1]*ay[i]-az[i]+X[2])/math.sqrt(X[0]*X[0]+X[1]*X[1]+1)
    i+=1


#print("PV =", max(pvbase)-min(pvbase))

#print('平面拟合结果为：z = %.8f * x + %.8f * y + %.8f' % (X[0], X[1], X[2]))

# -------------------------结果展示-------------------------------
fig1 = plt.figure()
ax1 = fig1.add_subplot(111, projection='3d')
ax1.set_xlabel("x")
ax1.set_ylabel("y")
ax1.set_zlabel("z")
ax1.scatter(x, y, z, c='r', marker='o')
x_p = np.linspace(min(ax), max(ax), 100)
y_p = np.linspace(min(ay), max(ay), 100)
x_p, y_p = np.meshgrid(x_p, y_p)
z_p = X[0] * x_p + X[1] * y_p + X[2]
ax1.plot_wireframe(x_p, y_p, z_p, rstride=10, cstride=10)

#保存图片
plt.savefig(filepath+filename+'.jpg', dpi=600)

# 结果写文件，依次为：文件名，A，B，C，D，PV；其中空间方程为Ax+By+Cz+D=0
workbook = openpyxl.load_workbook(filepath+'caldata.xlsx')
sheet = workbook['Sheet1']
data = [filename+'.xls',X[0][0], X[1][0],-1,X[2][0], max(pvbase)-min(pvbase)]
row = sheet.max_row + 1
for col, value in enumerate(data, start=1):
   sheet.cell(row=row, column=col).value = value
workbook.save(filepath+'caldata.xlsx')
workbook.close()
#窗口运行代码